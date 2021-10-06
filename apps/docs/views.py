from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import FileUploadParser
from .customResponse import CustomResponse
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.db.models import Count, Sum, OuterRef
from .serializer import DocCreateSerializer, LogSerializer, KategoriCreateSerializer, DetailDocCreateSerializer, DocDetailSerializer, BabSerializer, DocSerializer, BabSearchSerializer, DocTotalDetailSerializer, DocTotalDetailSerializerNoAnnotate
from .models import Document, Bab, DetailDocument, Log, Kategori, Babs
from datetime import datetime
import operator
import itertools
import re
from pathlib import Path
from notification.utils.CustomNotification import CustomNotification
from userinfo.models import User
from userinfo.views import authenticate_credentials
import openpyxl
import PyPDF2
import argparse
import subprocess
import os.path
import sys
from shutil import copyfile
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
import json


def checkIfRomanNumeral(numeral):
    """Controls that the userinput only contains valid roman numerals"""
    numeral = numeral.upper()
    validRomanNumerals = ["M", "D", "C", "L", "X", "V", "I"]
    for letters in numeral:
        if letters not in validRomanNumerals:
            # print("Sorry that is not a valid roman numeral")
            return False
    # romanToInt(numeral)
    return True


def compress(input_file_path, output_file_path, power=0):
    """Function to compress PDF via Ghostscript command line interface"""
    quality = {
        0: '/default',
        1: '/prepress',
        2: '/printer',
        3: '/ebook',
        4: '/screen'
    }

    # Basic controls
    # Check if valid path
    if not os.path.isfile(input_file_path):
        print("Error: invalid path for input PDF file")
        sys.exit(1)

    # Check if file is a PDF by extension
    if input_file_path.split('.')[-1].lower() != 'pdf':
        print("Error: input file is not a PDF")
        sys.exit(1)

    #print("Compress PDF...")
    initial_size = os.path.getsize(input_file_path)
    subprocess.call(['gs', '-sDEVICE=pdfwrite', '-dCompatibilityLevel=1.4',
                     '-dPDFSETTINGS={}'.format(quality[power]),
                     '-dNOPAUSE', '-dQUIET', '-dBATCH',
                     '-sOutputFile={}'.format(output_file_path),
                     input_file_path]
                    )
    final_size = os.path.getsize(output_file_path)
    ratio = 1 - (final_size / initial_size)
    #print("Compression by {0:.0%}.".format(ratio))
    #print("Final file size is {0:.1f}MB".format(final_size / 1000000))
    # print("Done.")


def getTextDoc(filename):
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
    return '\n'.join(fullText)


class DocsAPI(viewsets.ModelViewSet):
    def upload(self, request, format=None):
        token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        ret1, user = authenticate_credentials(token)
        if False == ret1 or None == user:
            return CustomResponse().badRequest(
                message='token invalid'
            )
        try:
            data_user = User.objects.get(id=request.data.get('user_id'))
        except User.DoesNotExist:
            return CustomResponse().base(
                success=False,
                message='user not found',
                status=200
            )
        if data_user.role == 'superadmin':
            # try:
            validSubBab = ["A.", "B.", "C.", "D.", "E.", "F.", "G.", "H.", "I.", "J.", "K.", "L.",
                           "M.", "N.", "O.", "P.", "Q.", "R.", "S.", "T.", "U.", "V.", "W.", "X.", "Y.", "Z."]
            pdf_file = request.FILES["pdf_file"]
            if not pdf_file:
                return Response.badRequest(message='No File Upload')
            kategori = request.data.get('kategori')
            level = request.data.get('level')

            # json_dict = {}
            # json_dict['kategori'] = kategori
            # json_dict['level'] = int(level)
            # serializer = DocCreateSerializer(data=json_dict)
            # if serializer.is_valid():
            #     serializer.save()
            # else:
            #     return CustomResponse().base(success=False, message=serializer.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            fs = FileSystemStorage(
                location=f'{settings.MEDIA_ROOT}/documents/',
                base_url=f'{settings.MEDIA_URL}/documents/'
            )
            filename = fs.save(pdf_file.name, pdf_file)
            file_path = fs.url(filename)

            # creating a pdf file object
            pdfFileObj = open(settings.BASE_DIR +
                              '/media/documents/'+filename, 'rb')
            # pdfFileObj = open('/Volumes/christopher2/datasintesa/digidoc_api/media/documents/Salinan K-BPK 1-2021_Juknis Pemeriksaan LKPD.pdf', 'rb')
            # pdfFileObj = open(file_path, 'rb')

            # content = ""

            # pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            # for i in range(0, pdfReader.getNumPages()):
            #     print(str(i))
            #     extractedText = pdfReader.getPage(14).extractText()
            #     content +=  extractedText + "\n"
            #     print(content.replace("\xa0", " "))
            #     content = " ".join(content.replace("\xa0", " ").strip().split())
            #     print(content.encode("ascii", "ignore"))
            #     return

            # content = " ".join(content.replace("\xa0", " ").strip().split())
            # print(content.encode("ascii", "ignore"))
            # return

            # creating a pdf reader object
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

            # printing number of pages in pdf file
            # print(pdfReader.numPages)
            # pageObj = pdfReader.getPage(6)
            # pagetext = pageObj.extractText()
            # print(pagetext)
            numPages = pdfReader.numPages
            total_page = 0
            stop = False
            last_pagecontens = 0
            list_contens = []
            # creating a page object
            for i in range(pdfReader.numPages):
                pageObj = pdfReader.getPage(i)

            #     # extracting text from page
                # and 'daftar isian' not in pageObj.extractText().lower() and 'daftar isi ' not in pageObj.extractText().lower():
                if 'DAFTAR ISI' in pageObj.extractText() and 'BAB' in pageObj.extractText():
                    pagetext = pageObj.extractText()
                    listtext = pagetext.splitlines()
                    lst_daftarisi = [i for i, x in enumerate(
                        listtext) if 'DAFTAR ISI' in x]
                    if len(lst_daftarisi) > 1:
                        listtext = listtext[lst_daftarisi[-1]:]
                    else:
                        listtext = listtext[lst_daftarisi[0]:]
                    listtext = list(filter(lambda x: '..' not in x, listtext))
                    listtext = list(filter(lambda x: x != ' ', listtext))
                    listtext = list(filter(lambda x: x != '', listtext))
                    #list_contens = []
                    list_items = []
                    items = ''
                    # for a in listtext:
                    for a in range(len(listtext)):
                        item = listtext[a]
                        # item = " ".join(item.replace("\xa0", " ").strip().split())
                        try:
                            if item[0] == '.':
                                item = item[1:]
                            item = int(item.strip())
                            for w in list_items:
                                items = items + listtext[w]
                            if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                list_contens.append(items)
                            items = ''
                            list_items = []
                        except:
                            if not checkIfRomanNumeral(item.strip()):
                                list_items.append(a)
                            else:
                                for w in list_items:
                                    items = items + listtext[w]
                                if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                    list_contens.append(items)
                                items = ''
                                list_items = []
                    # print(list_contens)
                    print('===================')
                    total_page += 1
                    start_new = i+1
                    for x in range(start_new, start_new+10):
                        print('page :', x)
                        last_pagecontens = x
                        pageObj = pdfReader.getPage(x)
                        pagetext = pageObj.extractText()
                        if 'daftar isi' in pagetext.lower() or 'BAB' in pagetext:
                            pagetext = pagetext.splitlines()
                            pagetext = list(
                                filter(lambda x: '..' not in x, pagetext))
                            pagetext = list(
                                filter(lambda x: x != ' ', pagetext))
                            pagetext = list(
                                filter(lambda x: x != '', pagetext))
                            items = ''
                            list_items = []
                            for a in range(len(pagetext)):
                                item = pagetext[a]
                                # item = " ".join(item.replace("\xa0", " ").strip().split())
                                try:
                                    if item[0] == '.':
                                        item = item[1:]
                                    item = int(item.strip())
                                    for w in list_items:
                                        items = items + pagetext[w]
                                    if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                        list_contens.append(items)
                                    items = ''
                                    list_items = []
                                except:
                                    if not checkIfRomanNumeral(item.strip()):
                                        list_items.append(a)
                                    else:
                                        for w in list_items:
                                            items = items + pagetext[w]
                                        if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                            list_contens.append(items)
                                        items = ''
                                        list_items = []
                            # print(list_contens)
                            print('===================')
                            total_page += 1
                        else:
                            stop = True
                            break
                if stop:
                    break

            if len(list_contens) == 0:
                return CustomResponse().base(
                    success=False,
                    message='file pdf tidak sesuai',
                    status=204
                )
            json_dict = {}
            json_dict['kategori'] = kategori
            json_dict['level'] = int(level)
            serializer = DocCreateSerializer(data=json_dict)
            if serializer.is_valid():
                serializer.save()
            else:
                return CustomResponse().base(success=False, message=serializer.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            json_dict = {}
            str_contens = '|'.join(map(str, list_contens))
            json_dict['contents'] = str_contens
            json_dict['doc_file'] = '/media/documents/'+filename
            json_dict['title'] = request.data.get('title')
            json_dict['pages'] = numPages
            json_dict['document'] = serializer.data['id']

            serializerDetail = DetailDocCreateSerializer(data=json_dict)
            if serializerDetail.is_valid():
                serializerDetail.save()
            else:
                return CustomResponse().base(success=False, message=serializerDetail.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            lst_bab = [i for i, x in enumerate(list_contens) if 'BAB' in x]
            # for x in lst_bab:
            bab = ''
            data_detaildoc = DetailDocument.objects.get(
                id=serializerDetail.data['id'])

            for x in range(lst_bab[0], len(list_contens)):
                lst_ = list_contens[x].split(' ')
                if 'BAB' in list_contens[x]:
                    if len(lst_) > 1:
                        lst_ = lst_[0]+' '+lst_[1]
                    else:
                        lst_ = lst_[0]
                else:
                    if len(lst_) > 1:
                        lst_ = ' '.join(lst_[1:])
                        lst_ = lst_.strip()
                    else:
                        lst_ = lst_[0]
                looping = 0
                for i in range(last_pagecontens, pdfReader.numPages):
                    # if looping>=10:
                    #     last_pagecontens=i
                    #     looping=0
                    #     break
                    pageObj = pdfReader.getPage(i)
                    pagetext = pageObj.extractText()
                    pagetext1 = pagetext.splitlines()
                    pagetext1 = [x.strip() for x in pagetext1]

                    found = False
                    # if 'BAB' in lst_:
                    if lst_ in pagetext1:
                        found = True
                    # else:
                    if not found:
                        str_pagetext1 = ''.join(map(str, pagetext1))
                        str_pagetext1 = "".join(str_pagetext1.replace(
                            "\xa0", "").replace(' ', '').strip().split())
                        find = lst_.split(' ')
                        find = ''.join(map(str, find))
                        if find in str_pagetext1:
                            found = True

                    if found:
                        # data_detaildoc = DetailDocument.objects.get(id=serializerDetail.data['id'])
                        # output pdf file name
                        # outputpdf = settings.BASE_DIR+'/media/documents/' + \
                        #     str(data_detaildoc.id)+'_'+str(i) + '.pdf'
                        if 'BAB' in list_contens[x]:
                            last_pagecontens_ = i
                            bab = lst_[0].upper()+lst_[1:3].lower() + \
                                ' '+lst_[4:]
                            pagetext = " ".join(pagetext.replace(
                                "\xa0", " ").strip().split())
                            data_bab = Bab(
                                bab=list_contens[x],
                                sub_bab='',
                                text=pagetext,
                                text1=str_pagetext1,
                                page=str(i),
                                path='/media/documents/' +
                                str(data_detaildoc.id)+'_'+str(i) + '.pdf',
                                detaildocument=data_detaildoc
                            )
                        else:
                            last_pagecontens = i
                            pagetext = " ".join(pagetext.replace(
                                "\xa0", " ").strip().split())
                            data_bab = Bab(
                                bab=bab,
                                sub_bab=list_contens[x],
                                text=pagetext,
                                text1=str_pagetext1,
                                page=str(i),
                                path='/media/documents/' +
                                str(data_detaildoc.id)+'_'+str(i) + '.pdf',
                                detaildocument=data_detaildoc
                            )

                        data_bab.save()
                        # looping+=1
                        # # creating pdf writer object for (i+1)th split
                        # pdfWriter = PyPDF2.PdfFileWriter()
                        # pdfWriter.addPage(pdfReader.getPage(i))

                        # # # writing split pdf pages to pdf file
                        # with open(outputpdf, "wb") as f:
                        #     pdfWriter.write(f)

                        break
                    if i == pdfReader.numPages:
                        last_pagecontens = last_pagecontens_
                # last_pagecontens+=1

            for ii in range(pdfReader.numPages):
                outputpdf = settings.BASE_DIR+'/media/documents/' + \
                    str(data_detaildoc.id)+'_'+str(ii) + '.pdf'

                pdfWriter = PyPDF2.PdfFileWriter()
                pdfWriter.addPage(pdfReader.getPage(ii))

                # writing split pdf pages to pdf file
                with open(outputpdf, "wb") as f:
                    pdfWriter.write(f)
            print('total_page', total_page)
            # closing the pdf file object
            pdfFileObj.close()

            # data_doc = Document.objects.get(id=serializer.data['id'])
            data_log = Log(
                action='upload',
                user=data_user,
                document=data_detaildoc
            )
            data_log.save()

            data_user_pemeriksa = User.objects.all()  # .exclude(role='superadmin')

            for du in data_user_pemeriksa:
                notif = CustomNotification()
                notif.create(to=du.id, from_=request.data.get('user_id'), type='upload',
                             title='Upload Document succeed', message='Dokumen baru telah terupload', push_message='Ada pesan baru')
            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        else:
            # return Response.badRequest(message='access denied')
            return CustomResponse().badRequest(
                message='access denied'
            )
        # except Exception as e:
        #    return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def upgradeversion(self, request, format=None):
        token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        ret1, user = authenticate_credentials(token)
        if False == ret1 or None == user:
            return CustomResponse().badRequest(
                message='token invalid'
            )
        try:
            data_user = User.objects.get(id=request.data.get('user_id'))
        except User.DoesNotExist:
            return CustomResponse().base(
                success=False,
                message='user not found',
                status=200
            )
        if data_user.role == 'superadmin':
            # try:
            validSubBab = ["A.", "B.", "C.", "D.", "E.", "F.", "G.", "H.", "I.", "J.", "K.", "L.",
                           "M.", "N.", "O.", "P.", "Q.", "R.", "S.", "T.", "U.", "V.", "W.", "X.", "Y.", "Z."]
            pdf_file = request.FILES["pdf_file"]
            if not pdf_file:
                return Response.badRequest(message='No File Upload')

            data_doc = Document.objects.get(id=request.data.get('id'))

            fs = FileSystemStorage(
                location=f'{settings.MEDIA_ROOT}/documents/',
                base_url=f'{settings.MEDIA_URL}/documents/'
            )
            filename = fs.save(pdf_file.name, pdf_file)
            file_path = fs.url(filename)

            # creating a pdf file object
            pdfFileObj = open(settings.BASE_DIR +
                              '/media/documents/'+filename, 'rb')

            # creating a pdf reader object
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

            numPages = pdfReader.numPages
            total_page = 0
            stop = False
            last_pagecontens = 0
            list_contens = []
            # creating a page object
            for i in range(pdfReader.numPages):
                pageObj = pdfReader.getPage(i)

            #     # extracting text from page
                # and 'daftar isian' not in pageObj.extractText().lower() and 'daftar isi ' not in pageObj.extractText().lower():
                if 'DAFTAR ISI' in pageObj.extractText() and 'BAB' in pageObj.extractText():
                    pagetext = pageObj.extractText()
                    listtext = pagetext.splitlines()
                    lst_daftarisi = [i for i, x in enumerate(
                        listtext) if 'DAFTAR ISI' in x]
                    if len(lst_daftarisi) > 1:
                        listtext = listtext[lst_daftarisi[-1]:]
                    else:
                        listtext = listtext[lst_daftarisi[0]:]
                    listtext = list(filter(lambda x: '..' not in x, listtext))
                    listtext = list(filter(lambda x: x != ' ', listtext))
                    listtext = list(filter(lambda x: x != '', listtext))
                    #list_contens = []
                    list_items = []
                    items = ''
                    # for a in listtext:
                    for a in range(len(listtext)):
                        item = listtext[a]
                        item = " ".join(item.replace(
                            "\xa0", " ").strip().split())
                        try:
                            if item[0] == '.':
                                item = item[1:]
                            item = int(item.strip())
                            for w in list_items:
                                items = items + listtext[w]
                            if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                list_contens.append(items)
                            items = ''
                            list_items = []
                        except:
                            if not checkIfRomanNumeral(item.strip()):
                                list_items.append(a)
                            else:
                                for w in list_items:
                                    items = items + listtext[w]
                                if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                    list_contens.append(items)
                                items = ''
                                list_items = []
                    # print(list_contens)
                    print('===================')
                    total_page += 1
                    start_new = i+1
                    for x in range(start_new, start_new+10):
                        print('page :', x)
                        last_pagecontens = x
                        pageObj = pdfReader.getPage(x)
                        pagetext = pageObj.extractText()
                        if 'daftar isi' in pagetext.lower():
                            pagetext = pagetext.splitlines()
                            pagetext = list(
                                filter(lambda x: '..' not in x, pagetext))
                            pagetext = list(
                                filter(lambda x: x != ' ', pagetext))
                            pagetext = list(
                                filter(lambda x: x != '', pagetext))
                            items = ''
                            list_items = []
                            for a in range(len(pagetext)):
                                item = pagetext[a]
                                item = " ".join(item.replace(
                                    "\xa0", " ").strip().split())
                                try:
                                    if item[0] == '.':
                                        item = item[1:]
                                    item = int(item.strip())
                                    for w in list_items:
                                        items = items + pagetext[w]
                                    if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                        list_contens.append(items)
                                    items = ''
                                    list_items = []
                                except:
                                    if not checkIfRomanNumeral(item.strip()):
                                        list_items.append(a)
                                    else:
                                        for w in list_items:
                                            items = items + pagetext[w]
                                        if 'BAB' == items[0:3] or items[0:2] in validSubBab:
                                            list_contens.append(items)
                                        items = ''
                                        list_items = []
                            # print(list_contens)
                            print('===================')
                            total_page += 1
                        else:
                            stop = True
                            break
                if stop:
                    break

            if len(list_contens) == 0:
                return CustomResponse().base(
                    success=False,
                    message='upload failed',
                    status=200
                )
            data_detaildoc = DetailDocument.objects.filter(
                document=data_doc.id).order_by('-version').first()
            if not data_detaildoc:
                return Response().base(
                    success=False,
                    message='detail document tidak ada',
                    status=200
                )

            str_contens = '|'.join(map(str, list_contens))
            data_detaildoc_ = DetailDocument(
                version=str(int(data_detaildoc.version)+1),
                contents=str_contens,
                doc_file='/media/documents/'+filename,
                title=request.data.get('title'),
                pages=numPages,
                document=data_doc
            )
            data_detaildoc_.save()
            lst_bab = [i for i, x in enumerate(list_contens) if 'BAB' in x]
            bab = ''
            for x in range(lst_bab[0], len(list_contens)):
                lst_ = list_contens[x].split(' ')
                if 'BAB' in list_contens[x]:
                    if len(lst_) > 1:
                        lst_ = lst_[0]+' '+lst_[1]
                    else:
                        lst_ = lst_[0]
                else:
                    if len(lst_) > 1:
                        lst_ = ' '.join(lst_[1:])
                        lst_ = lst_.strip()
                    else:
                        lst_ = lst_[0]
                looping = 0
                for i in range(last_pagecontens, pdfReader.numPages):
                    # if looping>=10:
                    #     last_pagecontens=i
                    #     looping=0
                    #     break
                    pageObj = pdfReader.getPage(i)
                    pagetext = pageObj.extractText()
                    pagetext1 = pagetext.splitlines()
                    pagetext1 = [x.strip() for x in pagetext1]

                    found = False
                    # if 'BAB' in lst_:
                    if lst_ in pagetext1:
                        str_pagetext1 = ''.join(map(str, pagetext1))
                        str_pagetext1 = "".join(str_pagetext1.replace(
                            "\xa0", "").replace(' ', '').strip().split())
                        found = True
                    # else:
                    if not found:
                        str_pagetext1 = ''.join(map(str, pagetext1))
                        str_pagetext1 = "".join(str_pagetext1.replace(
                            "\xa0", "").replace(' ', '').strip().split())
                        find = lst_.split(' ')
                        find = ''.join(map(str, find))
                        if find in str_pagetext1:
                            found = True
                    if found:
                        # data_detaildoc = DetailDocument.objects.get(id=serializerDetail.data['id'])
                        if 'BAB' in list_contens[x]:
                            last_pagecontens_ = i
                            bab = lst_[0].upper()+lst_[1:3].lower() + \
                                ' '+lst_[4:]
                            pagetext = " ".join(pagetext.replace(
                                "\xa0", " ").strip().split())
                            data_bab = Bab(
                                bab=list_contens[x],
                                sub_bab='',
                                text=pagetext,
                                text1=str_pagetext1,
                                page=str(i),
                                path='/media/documents/' +
                                str(data_detaildoc_.id)+'_'+str(i) + '.pdf',
                                detaildocument=data_detaildoc_
                            )
                        else:
                            last_pagecontens = i
                            pagetext = " ".join(pagetext.replace(
                                "\xa0", " ").strip().split())
                            data_bab = Bab(
                                bab=bab,
                                sub_bab=list_contens[x],
                                text=pagetext,
                                text1=str_pagetext1,
                                page=str(i), path='/media/documents/' +
                                str(data_detaildoc_.id)+'_'+str(i) + '.pdf',
                                detaildocument=data_detaildoc_
                            )

                        data_bab.save()
                        # looping+=1
                        break
                    if i == pdfReader.numPages:
                        last_pagecontens = last_pagecontens_
                # last_pagecontens+=1

            for ii in range(pdfReader.numPages):
                outputpdf = settings.BASE_DIR+'/media/documents/' + \
                    str(data_detaildoc_.id)+'_'+str(ii) + '.pdf'

                pdfWriter = PyPDF2.PdfFileWriter()
                pdfWriter.addPage(pdfReader.getPage(ii))

                # writing split pdf pages to pdf file
                with open(outputpdf, "wb") as f:
                    pdfWriter.write(f)

            print('total_page', total_page)
            # closing the pdf file object
            pdfFileObj.close()

            # data_doc = Document.objects.get(id=serializer.data['id'])
            data_log = Log(
                action='upgrade version',
                user=data_user,
                document=data_detaildoc_
            )
            data_log.save()

            data_user_pemeriksa = User.objects.all()  # .exclude(role='superadmin')

            for du in data_user_pemeriksa:
                notif = CustomNotification()
                notif.create(to=du.id, from_=request.data.get('user_id'), type='update version',
                             title='Update version Document succeed', message='Versi terbaru ' +
                             data_detaildoc_.title+' telah terupload', push_message='Ada pesan baru')
            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        else:
            return CustomResponse().badRequest(
                message='access denied'
            )
        # except Exception as e:
        #    return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def upload_doc(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()
            if data_user.role == 'superadmin':
                kategori = int(request.data.get('kategori'))
                level = request.data.get('level')
                daftar_isi = request.FILES["daftar_isi"]
                if not daftar_isi:
                    return Response.badRequest(message='No File daftar isi')

                pdf_file = request.FILES["pdf_file"]
                if not pdf_file:
                    return Response.badRequest(message='No File Upload')
                # you may put validations here to check extension or file size

                fs_daftarisi = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/documents/',
                    base_url=f'{settings.MEDIA_URL}/documents/'
                )
                filename_daftarisi = fs_daftarisi.save(
                    daftar_isi.name, daftar_isi)
                file_path_daftarisi = fs_daftarisi.url(filename_daftarisi)

                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/documents/',
                    base_url=f'{settings.MEDIA_URL}/documents/'
                )
                filename = fs.save(pdf_file.name, pdf_file)
                file_path = fs.url(filename)

                wb = openpyxl.load_workbook(daftar_isi)

                # getting a particular sheet by name out of many sheets
                worksheet = wb["Sheet2"]

                # creating a pdf file object
                pdfFileObj = open(settings.BASE_DIR +
                                  '/media/documents/'+filename, 'rb')

                # creating a pdf reader object
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

                numPages = pdfReader.numPages

                json_dict = {}
                json_dict['kategori'] = kategori
                json_dict['level'] = int(level)
                serializer = DocCreateSerializer(data=json_dict)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return CustomResponse().base(success=False, message=serializer.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

                json_dict = {}
                str_contens = '|'
                json_dict['contents'] = str_contens
                json_dict['doc_file'] = '/media/documents/'+filename
                json_dict['title'] = request.data.get('title')
                json_dict['pages'] = numPages
                json_dict['document'] = serializer.data['id']

                serializerDetail = DetailDocCreateSerializer(data=json_dict)
                if serializerDetail.is_valid():
                    serializerDetail.save()
                else:
                    return CustomResponse().base(success=False, message=serializerDetail.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

                data_detaildoc = DetailDocument.objects.get(
                    id=serializerDetail.data['id'])
                bab = ''
                list_contens = []
                multi_pages = []
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    if str(row[0].value) == 'None':
                        break
                    if str(row[0].value) == 'BAB':
                        continue
                    if str(row[0].value) == 'None' or str(row[1].value) == 'None' or str(row[2].value) == 'None' or str(row[3].value) == 'None':
                        data_detaildoc.status = 'deactivate'
                        data_detaildoc.title = data_detaildoc.title+'-'+data_detaildoc.id
                        data_detaildoc.update_date = datetime.now()
                        data_detaildoc.save()
                        return CustomResponse().badRequest(
                            message="excel's columns can't null"
                        )
                    if int(str(row[2].value))-1 > numPages or int(str(row[3].value))-1 > numPages:
                        return CustomResponse().badRequest(
                            message="start page or end page can't more than total page"
                        )
                    search_text = ''
                    search_text1 = ''
                    is_bab = True
                    if str(row[1].value).strip() == '-':
                        search_text1 = str(row[0].value).strip()
                        list_contens.append('|'+search_text1)
                        is_bab = True
                    else:
                        search_text1 = str(row[1].value).strip()
                        list_contens.append('#'+search_text1)
                        is_bab = False
                    search_text = search_text1.replace(' ', '')

                    multi_page = []
                    start_page = int(str(row[2].value).strip())-1
                    end_page = int(str(row[3].value).strip())-1
                    pdf_name = str(start_page)+'_'+str(end_page)
                    if end_page > start_page:
                        multi_page.append(start_page)
                        multi_page.append(end_page)
                    if len(multi_page) > 1:
                        pages = multi_page
                        multi_pages.append(pages)
                    else:
                        pages = []
                        pages.append(start_page)
                        pages.append(end_page)
                        pdf_name = str(start_page)

                    # creating a page object
                    # for i in range(start_page,numPages):
                    lst_pagetext = []
                    lst_str_pagetext1 = []
                    # if len(pages)==1:
                    #     pageObj = pdfReader.getPage(pages[0])
                    #     #pdf_name.append(pages[i])
                    # #     # extracting text from page
                    #     pagetext = pageObj.extractText()
                    #     pagetext1 = pagetext.splitlines()
                    #     pagetext1 = [x.strip() for x in pagetext1]

                    #     str_pagetext1 = ''.join(map(str, pagetext1))
                    #     str_pagetext1 = "".join(str_pagetext1.replace(
                    #         "\xa0", "").replace(' ', '').strip().split())
                    #     lst_str_pagetext1.append(str_pagetext1)
                    #     #if search_text in str_pagetext1:
                    #     pagetext = " ".join(pagetext.replace(
                    #                 "\xa0", " ").strip().split())
                    #     lst_pagetext.append(pagetext)
                    # else:
                    # for i in range(len(pages)):
                    for i in range(pages[0], pages[1]+1):
                        pageObj = pdfReader.getPage(i)
                        # pdf_name.append(pages[i])
                    #     # extracting text from page
                        pagetext = pageObj.extractText()
                        pagetext1 = pagetext.splitlines()
                        pagetext1 = [x.strip() for x in pagetext1]

                        str_pagetext1 = ''.join(map(str, pagetext1))
                        str_pagetext1 = "".join(str_pagetext1.replace(
                            "\xa0", "").replace(' ', '').strip().split())
                        lst_str_pagetext1.append(str_pagetext1)
                        # if search_text in str_pagetext1:
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        lst_pagetext.append(pagetext)

                    #pdf_name='_'.join(map(str, pdf_name))
                    # if 'BAB' in search_text:
                    if is_bab:
                        #last_pagecontens_ = i
                        bab = search_text1
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        data_bab = Bab(
                            bab=search_text1,
                            sub_bab='',
                            text='\n'.join(map(str, lst_pagetext)),
                            text1='\n'.join(map(str, lst_str_pagetext1)),
                            page=str(start_page),
                            endpage=str(end_page),
                            path='/media/documents/' + str(serializer.data['id'])+'_' +
                            str(data_detaildoc.id)+'_'+pdf_name + '.pdf',
                            detaildocument=data_detaildoc
                        )
                    else:
                        #last_pagecontens = i
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        data_bab = Bab(
                            bab=bab,
                            sub_bab=search_text1,
                            text='\n'.join(map(str, lst_pagetext)),
                            text1='\n'.join(map(str, lst_str_pagetext1)),
                            page=str(start_page),
                            endpage=str(end_page),
                            path='/media/documents/' + str(serializer.data['id'])+'_' +
                            str(data_detaildoc.id)+'_'+pdf_name + '.pdf',
                            detaildocument=data_detaildoc
                        )

                    data_bab.save()
                    # break

                data_detaildoc.contents = ''.join(map(str, list_contens))
                data_detaildoc.save()
                for ii in range(numPages):
                    outputpdf = settings.BASE_DIR+'/media/documents/' + str(serializer.data['id'])+'_' + \
                        str(data_detaildoc.id)+'_'+str(ii) + '.pdf'

                    pdfWriter = PyPDF2.PdfFileWriter()
                    # pdfWriter.addPage(pdfReader.getPage(ii))
                    page = pdfReader.getPage(ii)
                    page.compressContentStreams()
                    pdfWriter.addPage(page)

                    # writing split pdf pages to pdf file
                    with open(outputpdf, "wb") as f:
                        pdfWriter.write(f)

                    # compress(outputpdf, settings.BASE_DIR+'/media/documents/' + str(serializer.data['id'])+'_' +
                    #          str(data_detaildoc.id)+'_'+str(ii) + '.pdf', power=0)
                    # os.remove(outputpdf)
                    # subprocess.run('ps2pdf '+outputpdf+' '+settings.BASE_DIR+'/media/documents/' + \
                    #     str(data_detaildoc.id)+'_'+str(ii) + '.pdf', shell=True)
                    # subprocess.run('rm -R '+outputpdf, shell=True)
                    pageObj = pdfReader.getPage(ii)
                    pagetext = pageObj.extractText()
                    pagetext1 = pagetext.splitlines()
                    pagetext1 = [x.strip() for x in pagetext1]

                    pagetext = " ".join(pagetext.replace(
                        "\xa0", " ").strip().split())

                    str_pagetext1 = ''.join(map(str, pagetext1))
                    str_pagetext1 = "".join(str_pagetext1.replace(
                        "\xa0", "").replace(' ', '').strip().split())
                    data_babs = Babs(
                        text=pagetext,
                        text1=str_pagetext1,
                        page=str(ii),
                        path='/media/documents/' + str(serializer.data['id'])+'_' +
                        str(data_detaildoc.id)+'_'+str(ii) + '.pdf',
                        detaildocument=data_detaildoc
                    )

                    data_babs.save()

                if len(multi_pages) > 0:
                    #pdfMerger = PyPDF2.PdfFileMerger()
                    for s in multi_pages:
                        pdfMerger = PyPDF2.PdfFileMerger()
                        for ss in range(s[0], s[1]+1):
                            ffile = settings.BASE_DIR+'/media/documents/' + str(serializer.data['id'])+'_' +\
                                str(data_detaildoc.id)+'_'+str(ss) + '.pdf'
                            pdfMerger.append(ffile)
                        pdfMerger.write(settings.BASE_DIR+'/media/documents/' + str(serializer.data['id'])+'_' +
                                        str(data_detaildoc.id)+'_'+'_'.join(map(str, s))+'.pdf')
                # closing the pdf file object
                pdfFileObj.close()

                data_log = Log(
                    action='upload',
                    user=data_user,
                    document=data_detaildoc
                )
                data_log.save()

                data_user_pemeriksa = User.objects.all()  # .exclude(role='superadmin')
                json_doc = {}
                json_doc['doc_id'] = serializer.data['id']
                json_doc['detaildoc_id'] = data_detaildoc.id
                json_doc['judul'] = data_detaildoc.title
                json_doc['kategori'] = serializer.data['kategori']
                json_doc['versi'] = data_detaildoc.version

                for du in data_user_pemeriksa:
                    notif = CustomNotification()
                    notif.create(to=du.id, from_=request.data.get('user_id'), type='upload',
                                 title='Upload Document succeed', message='Dokumen baru telah terupload. Dokumen ' +
                                 request.data.get('title')+' Telah Terupload', push_message='Ada pesan baru', 
                                 detail=json.dumps(json_doc))
                return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
            else:
                return CustomResponse().badRequest(
                    message='access denied'
                )
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def upgradeversion_doc(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()
            if data_user.role == 'superadmin':
                daftar_isi = request.FILES["daftar_isi"]
                if not daftar_isi:
                    return Response.badRequest(message='No File daftar isi')

                pdf_file = request.FILES["pdf_file"]
                if not pdf_file:
                    return Response.badRequest(message='No File Upload')

                data_doc = Document.objects.get(id=request.data.get('id'))

                # you may put validations here to check extension or file size
                fs_daftarisi = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/documents/',
                    base_url=f'{settings.MEDIA_URL}/documents/'
                )
                filename_daftarisi = fs_daftarisi.save(
                    daftar_isi.name, daftar_isi)
                file_path_daftarisi = fs_daftarisi.url(filename_daftarisi)

                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/documents/',
                    base_url=f'{settings.MEDIA_URL}/documents/'
                )
                filename = fs.save(pdf_file.name, pdf_file)
                file_path = fs.url(filename)

                wb = openpyxl.load_workbook(daftar_isi)

                # getting a particular sheet by name out of many sheets
                worksheet = wb["Sheet2"]

                # creating a pdf file object
                pdfFileObj = open(settings.BASE_DIR +
                                  '/media/documents/'+filename, 'rb')

                # creating a pdf reader object
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

                numPages = pdfReader.numPages

                data_detaildoc = DetailDocument.objects.filter(
                    document=data_doc.id).order_by('-version').first()
                if not data_detaildoc:
                    return Response().base(
                        success=False,
                        message='detail document tidak ada',
                        status=200
                    )

                str_contens = '|'
                data_detaildoc_ = DetailDocument(
                    version=str(int(data_detaildoc.version)+1),
                    contents=str_contens,
                    doc_file='/media/documents/'+filename,
                    title=request.data.get('title'),
                    pages=numPages,
                    document=data_doc
                )
                data_detaildoc_.save()

                bab = ''
                list_contens = []
                multi_pages = []
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    if str(row[0].value) == 'None':
                        break
                    if str(row[0].value) == 'BAB':
                        continue
                    if str(row[0].value) == 'None' or str(row[1].value) == 'None' or str(row[2].value) == 'None' or str(row[3].value) == 'None':
                        data_detaildoc_.status = 'deactivate'
                        data_detaildoc.title = data_detaildoc.title+'-'+data_detaildoc.id
                        data_detaildoc_.update_date = datetime.now()
                        data_detaildoc_.save()
                        return CustomResponse().badRequest(
                            message="excel's columns can't null"
                        )
                    search_text = ''
                    search_text1 = ''
                    is_bab = True
                    if str(row[1].value).strip() == '-':
                        search_text1 = str(row[0].value).strip()
                        list_contens.append('|'+search_text1)
                        is_bab = True
                    else:
                        search_text1 = str(row[1].value).strip()
                        list_contens.append('#'+search_text1)
                        is_bab = False
                    search_text = search_text1.replace(' ', '')

                    multi_page = []
                    start_page = int(str(row[2].value).strip())-1
                    end_page = int(str(row[3].value).strip())-1
                    pdf_name = str(start_page)+'_'+str(end_page)
                    if end_page > start_page:
                        multi_page.append(start_page)
                        multi_page.append(end_page)
                    if len(multi_page) > 1:
                        pages = multi_page
                        multi_pages.append(pages)
                    else:
                        pages = []
                        pages.append(start_page)
                        pages.append(end_page)
                        pdf_name = str(start_page)

                    # creating a page object
                    # for i in range(start_page,numPages):
                    lst_pagetext = []
                    lst_str_pagetext1 = []
                    for i in range(pages[0], pages[1]+1):
                        pageObj = pdfReader.getPage(i)
                        # pdf_name.append(pages[i])
                    #     # extracting text from page
                        pagetext = pageObj.extractText()
                        pagetext1 = pagetext.splitlines()
                        pagetext1 = [x.strip() for x in pagetext1]

                        str_pagetext1 = ''.join(map(str, pagetext1))
                        str_pagetext1 = "".join(str_pagetext1.replace(
                            "\xa0", "").replace(' ', '').strip().split())
                        lst_str_pagetext1.append(str_pagetext1)
                        # if search_text in str_pagetext1:
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        lst_pagetext.append(pagetext)

                    #pdf_name='_'.join(map(str, pdf_name))
                    # if 'BAB' in search_text:
                    if is_bab:
                        #last_pagecontens_ = i
                        bab = search_text1
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        data_bab = Bab(
                            bab=search_text1,
                            sub_bab='',
                            text='\n'.join(map(str, lst_pagetext)),
                            text1='\n'.join(map(str, lst_str_pagetext1)),
                            page=str(start_page),
                            endpage=str(end_page),
                            path='/media/documents/' + str(data_doc.id)+'_' +
                            str(data_detaildoc_.id)+'_'+pdf_name + '.pdf',
                            detaildocument=data_detaildoc_
                        )
                    else:
                        #last_pagecontens = i
                        pagetext = " ".join(pagetext.replace(
                            "\xa0", " ").strip().split())
                        data_bab = Bab(
                            bab=bab,
                            sub_bab=search_text1,
                            text='\n'.join(map(str, lst_pagetext)),
                            text1='\n'.join(map(str, lst_str_pagetext1)),
                            page=str(start_page),
                            endpage=str(end_page),
                            path='/media/documents/' + str(data_doc.id)+'_' +
                            str(data_detaildoc_.id)+'_'+pdf_name + '.pdf',
                            detaildocument=data_detaildoc_
                        )

                    data_bab.save()
                    # break

                data_detaildoc_.contents = ''.join(map(str, list_contens))
                data_detaildoc_.save()
                for ii in range(numPages):
                    outputpdf = settings.BASE_DIR+'/media/documents/' + str(data_doc.id)+'_' +\
                        str(data_detaildoc_.id)+'_'+str(ii) + '.pdf'

                    pdfWriter = PyPDF2.PdfFileWriter()
                    # pdfWriter.addPage(pdfReader.getPage(ii))
                    page = pdfReader.getPage(ii)
                    page.compressContentStreams()
                    pdfWriter.addPage(page)

                    # writing split pdf pages to pdf file
                    with open(outputpdf, "wb") as f:
                        pdfWriter.write(f)

                    # compress(outputpdf, settings.BASE_DIR+'/media/documents/' + str(data_doc.id)+'_' +
                    #          str(data_detaildoc_.id)+'_'+str(ii) + '.pdf', power=0)
                    # os.remove(outputpdf)

                    pageObj = pdfReader.getPage(ii)
                    pagetext = pageObj.extractText()
                    pagetext1 = pagetext.splitlines()
                    pagetext1 = [x.strip() for x in pagetext1]

                    pagetext = " ".join(pagetext.replace(
                        "\xa0", " ").strip().split())

                    str_pagetext1 = ''.join(map(str, pagetext1))
                    str_pagetext1 = "".join(str_pagetext1.replace(
                        "\xa0", "").replace(' ', '').strip().split())
                    data_babs = Babs(
                        text=pagetext,
                        text1=str_pagetext1,
                        page=str(ii),
                        path='/media/documents/' + str(data_doc.id)+'_' +
                        str(data_detaildoc_.id)+'_'+str(ii) + '.pdf',
                        detaildocument=data_detaildoc_
                    )

                    data_babs.save()

                if len(multi_pages) > 0:
                    # pdfMerger = PyPDF2.PdfFileMerger()
                    for s in multi_pages:
                        pdfMerger = PyPDF2.PdfFileMerger()
                        for ss in range(s[0], s[1]+1):
                            ffile = settings.BASE_DIR+'/media/documents/' + str(data_doc.id)+'_' +\
                                str(data_detaildoc_.id)+'_'+str(ss) + '.pdf'
                            pdfMerger.append(ffile)
                        pdfMerger.write(settings.BASE_DIR+'/media/documents/' + str(data_doc.id)+'_' +
                                        str(data_detaildoc_.id)+'_'+'_'.join(map(str, s))+'.pdf')
                # closing the pdf file object
                pdfFileObj.close()

                data_log = Log(
                    action='upgrade version',
                    user=data_user,
                    document=data_detaildoc_
                )
                data_log.save()

                data_user_pemeriksa = User.objects.all()  # .exclude(role='superadmin')

                json_doc = {}
                json_doc['doc_id'] = data_doc.id
                json_doc['detaildoc_id'] = data_detaildoc_.id
                json_doc['judul'] = data_detaildoc_.title
                json_doc['kategori'] = data_doc.kategori.id
                json_doc['versi'] = data_detaildoc.version

                for du in data_user_pemeriksa:
                    notif = CustomNotification()
                    notif.create(to=du.id, from_=request.data.get('user_id'), type='update version',
                                 title='Update version Document succeed', message='Update Versi Baru Dokumen. Versi terbaru ' +
                                 data_detaildoc_.title+' telah terupload', push_message='Ada pesan baru',
                                 detail=json.dumps(json_doc))

                return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
            else:
                return CustomResponse().badRequest(
                    message='access denied'
                )
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def addlike1(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            data_doc = Document.objects.get(id=request.data.get('doc_id'))
            data_detaildoc = DetailDocument.objects.filter(
                document=data_doc.id).order_by('-version').first()
            list_likes = data_doc.likes.split(',')
            if request.data.get('user_id') in list_likes:
                # return CustomResponse().base(
                #     success=False,
                #     message='user liked already',
                #     status=200
                # )
                list_likes.remove(request.data.get('user_id'))
                data_doc.likes_count -= 1
            else:
                list_likes.append(request.data.get('user_id'))
                #data_doc.likes += ','+request.data.get('user_id')
                data_doc.likes_count += 1
            data_doc.likes = ','.join(map(str, list_likes))
            data_doc.update_date = datetime.now()
            data_doc.save()

            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()
            data_log = Log(
                action='like document',
                user=data_user,
                document=data_detaildoc
            )
            data_log.save()

            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def addlike(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            #data_doc = Document.objects.get(id=request.data.get('doc_id'))
            data_detaildoc = DetailDocument.objects.get(
                id=request.data.get('detaildoc_id'))
            # data_bab = Bab.objects.get(id=request.data.get('bab_id'))
            data_bab = Bab.objects.get(detaildocument=data_detaildoc.id, bab=request.data.get('bab'),sub_bab=request.data.get('subbab'))
            
            list_likes=[]
            list_likes_detaildoc=[]
            if data_bab.likes:
                list_likes = data_bab.likes.split(',')
            if request.data.get('user_id') in list_likes:
                list_likes.remove(request.data.get('user_id'))
                data_bab.likes_count -= 1
            else:
                list_likes.append(request.data.get('user_id'))
                data_bab.likes_count += 1
            data_bab.likes = ','.join(map(str, list_likes))
            data_bab.update_date = datetime.now()
            data_bab.save()

            data_bab_likes = Bab.objects.filter(detaildocument_id=request.data.get('detaildoc_id'), likes_count__gt=0)

            if data_bab_likes:
                total_like = 0

                for o in data_bab_likes:
                    total_like+=o.likes_count
                    list_likes_d = o.likes.split(',')
                    list_likes_detaildoc+=list_likes_d

                # if data_detaildoc.likes:
                #     list_likes_detaildoc = data_detaildoc.likes.split(',')
            
                # list_likes_detaildoc.append(request.data.get('user_id'))
                data_detaildoc.likes = ','.join(map(str, list_likes_detaildoc))
                data_detaildoc.likes_count = total_like
                data_detaildoc.save()

            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()
            data_log = Log(
                action='like bab/subbab',
                user=data_user,
                document=data_detaildoc
            )
            data_log.save()

            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def addview1(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            data_doc = Document.objects.get(id=request.data.get('doc_id'))
            data_detaildoc = DetailDocument.objects.filter(
                document=data_doc.id).order_by('-version').first()
            list_views = data_doc.views.split(',')
            if request.data.get('user_id') in list_views:
                list_views.remove(request.data.get('user_id'))
                data_doc.views_count -= 1
            else:
                list_views.append(request.data.get('user_id'))
                #data_doc.likes += ','+request.data.get('user_id')
                data_doc.views_count += 1
            data_doc.views = ','.join(map(str, list_views))
            data_doc.update_date = datetime.now()
            data_doc.save()

            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()

            data_log = Log(
                action='view document',
                user=data_user,
                document=data_detaildoc
            )
            data_log.save()

            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def addview(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            # data_doc = Document.objects.get(id=request.data.get('doc_id'))
            # data_detaildoc = DetailDocument.objects.filter(
            #     document=data_doc.id).order_by('-version').first()
            data_detaildoc = DetailDocument.objects.get(
                id=request.data.get('detaildoc_id'))
            list_views = []
            if data_detaildoc.views:
                list_views = data_detaildoc.views.split(',')
            #if request.data.get('user_id') in list_views:
                # list_views.remove(request.data.get('user_id'))
                # data_detaildoc.views_count -= 1
                # pass
            # else:
            list_views.append(request.data.get('user_id'))
            data_detaildoc.views_count += 1
            data_detaildoc.views = ','.join(map(str, list_views))
            data_detaildoc.update_date = datetime.now()
            data_detaildoc.save()

            try:
                data_user = User.objects.get(id=request.data.get('user_id'))
            except User.DoesNotExist:
                return CustomResponse().base(
                    success=False,
                    message='user not found',
                    status=200
                )
            data_user.current_log = datetime.now()
            data_user.save()
            data_log = Log(
                action='view document',
                user=data_user,
                document=data_detaildoc
            )
            data_log.save()

            return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        except Exception as e:
            return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def addopen(self, request, format=None):
        token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        ret1, user = authenticate_credentials(token)
        if False == ret1 or None == user:
            return CustomResponse().badRequest(
                message='token invalid'
            )
        # try:
        #data_doc = Document.objects.get(id=request.data.get('id'))
        data_detaildoc = DetailDocument.objects.get(
            id=request.data.get('detaildoc_id'))

        try:
            data_user = User.objects.get(id=request.data.get('user_id'))
        except User.DoesNotExist:
            return CustomResponse().base(
                success=False,
                message='user not found',
                status=200
            )
        data_user.current_log = datetime.now()
        data_user.save()
        data_log = Log(
            action='open document',
            user=data_user,
            document=data_detaildoc
        )
        data_log.save()

        return CustomResponse().base(values=[], status=status.HTTP_201_CREATED)
        # except Exception as e:
        #    return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def deactivate(self, request, format=None):
        token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        ret1, user = authenticate_credentials(token)
        if False == ret1 or None == user:
            return CustomResponse().badRequest(
                message='token invalid'
            )
        # try:
        #data_doc = Document.objects.get(id=request.data.get('id'))
        data_detaildoc = DetailDocument.objects.get(
            id=request.data.get('detaildoc_id'))
        data_detaildoc.status = 'deactivate'
        data_detaildoc.title = data_detaildoc.title+'-'+data_detaildoc.id
        data_detaildoc.update_date = datetime.now()
        data_detaildoc.save()

        try:
            data_user = User.objects.get(id=request.data.get('user_id'))
        except User.DoesNotExist:
            return CustomResponse().base(
                success=False,
                message='user not found',
                status=200
            )
        data_user.current_log = datetime.now()
        data_user.save()
        data_log = Log(
            action='delete document',
            user=data_user,
            document=data_detaildoc
        )
        data_log.save()

        return CustomResponse().base(values=data_detaildoc.serialize(), status=status.HTTP_201_CREATED)
        # except Exception as e:
        #    return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def editText(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            bab_id = request.data.get('bab_id')
            text = request.data.get('text')
            data_bab = Bab.objects.get(id=bab_id)
            data_bab.text = text
            data_bab.save()
            result = []
            result.append({
                "id": data_bab.id,
                "bab": data_bab.bab,
                "sub_bab": data_bab.sub_bab,
                "text": data_bab.text,
                "text1": data_bab.text1,
                "page": data_bab.page,
                "endpage": data_bab.endpage,
                "path": data_bab.path,
                "create_date": data_bab.create_date,
                "update_date": data_bab.update_date,
                "detaildoc_id": data_bab.detaildocument.id,
                "likes": data_bab.likes,
                "likes_count": data_bab.likes_count,
            })

            return CustomResponse.ok(values=result)
        except Exception as e:
           return CustomResponse().base(success=False, message=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DocGet(viewsets.ModelViewSet):
    def getDocument(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            category = request.query_params.get('kategori', None)
            if not category:
                return CustomResponse.badRequest('need query param "kategori"')
            data = Document.objects.filter(
                kategori=category).only("id").values_list('id')
            detail = DetailDocument.objects.filter(
                document__in=data, status='active')
            serializer = DocDetailSerializer(detail, many=True)

            def parseDaftarIsi(text: str):
                res = []
                bab = list(filter(lambda x: x, text.split('|')))
                for x in bab:
                    subBab = x.split('#')
                    res.append({
                        "bab": subBab.pop(0),
                        "list": subBab,
                    })
                return res

            result = []
            for x in serializer.data:
                result.append({
                    "doc_id": x['document'],
                    "id": x['id'],
                    "status": x['status'],
                    "kategori": category,
                    "pages": x['pages'],
                    "version": x['version'],
                    "title": x['title'],
                    "contents": parseDaftarIsi(x['contents'])
                })
            return CustomResponse.ok(values=result)
            # return CustomResponse().base(message=str(serializer.errors), status=500, success=False)
        except Document.DoesNotExist:
            return CustomResponse().base(message="document not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getDocumentText(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            detaildoc_id = request.query_params.get('detaildoc_id', None)
            if not detaildoc_id:
                return CustomResponse.badRequest('need query param "detaildoc_id"')
            babText = request.query_params.get('bab', None)
            if not babText:
                return CustomResponse.badRequest('need query param "bab"')
            subbabText = request.query_params.get('sub_bab', None)
            if not subbabText:
                return CustomResponse.badRequest('need query param "sub_bab"')

            data_bab = Bab.objects.filter(detaildocument=detaildoc_id,bab=babText,sub_bab=subbabText)
            babSerializer = BabSearchSerializer(data_bab, many=True)
            return CustomResponse.ok(values=babSerializer.data)
            # return CustomResponse().base(message=str(serializer.errors), status=500, success=False)
        except Document.DoesNotExist:
            return CustomResponse().base(message="document not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getDocumentsById(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            id = request.query_params.get('id', None)
            bab = request.query_params.get('bab', None)
            chap = request.query_params.get('chap', '')
            result = []
            if not id:
                return CustomResponse.badRequest('need query param "id"')
            if not bab:
                return CustomResponse.badRequest('need query param "bab"')

            if chap == '':
                BabAwal = Bab.objects.filter(
                    detaildocument=id, bab=bab, sub_bab='')
            else:
                BabAwal = Bab.objects.filter(
                    detaildocument=id, bab=bab, sub_bab=chap)
            serializer = BabSerializer(BabAwal, many=True)
            iddoc = serializer.data[0]['id']
            pageAwal = serializer.data[0]['page']
            pageAkhir = serializer.data[0]['endpage']
            result.append({'path': serializer.data[0]['path']})
            return CustomResponse.ok(values=result)
        except Document.DoesNotExist:
            return CustomResponse().base(message="document not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getDocumentById(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            id = request.query_params.get('id', None)
            if id:
                detail = DetailDocument.objects.filter(id=id, status='active')
            else:
                # detail = DetailDocument.objects.all()
                detail = DetailDocument.objects.filter(status='active')
            serializer = DocDetailSerializer(detail, many=True)

            def parseDaftarIsi(text: str):
                res = []
                bab = list(filter(lambda x: x, text.split('|')))
                for x in bab:
                    subBab = x.split('#')
                    res.append({
                        "bab": subBab.pop(0),
                        "list": subBab,
                    })
                return res

            result = []
            for x in serializer.data:
                doc = Document.objects.get(id=x['document'])
                kategori = Kategori.objects.get(id=doc.kategori.id)
                # parent_kategori = Kategori.objects.get(id=doc.kategori.id)
                # babLikes = Bab.objects.filter(
                #     detaildocument=x['id']).values_list('likes', flat=True)
                babLikes = Bab.objects.filter(
                    detaildocument=x['id'],likes_count__gt=0)
                # listLikes = list(filter(None, babLikes))
                # listLikes = [x.split(',') for x in listLikes]
                # listLikes = itertools.chain.from_iterable(listLikes)
                listLikes = []
                total_like=0
                for b in babLikes:
                    # if b.likes_count > 0:
                    json_dict = {}
                    json_dict['bab']=b.bab
                    json_dict['sub_bab']=b.sub_bab
                    json_dict['user_like'] = b.likes.split(',')
                    total_like +=len(json_dict['user_like'])
                    listLikes.append(json_dict)

                parent_id=''
                try:
                    parent_id = kategori.parent.id
                except:
                    pass
                result.append({
                    "doc_id": x['document'],
                    "id": x['id'],
                    "status": x['status'],
                    'pages': x['pages'],
                    "title": x['title'],
                    "id_kategori": kategori.id,
                    "kategori": kategori.name,
                    "parent_kategori": parent_id,
                    "versi": x['version'],
                    #"like": list(set(listLikes)),
                    # "like": list(listLikes),
                    "like": listLikes,
                    "total_like": total_like,
                    "view": x['views'].split(',') if x['views'] else [],
                    "file": x['doc_file'],
                    "contents": parseDaftarIsi(x['contents']),
                    "create_date": x['create_date']
                })
            return CustomResponse.ok(values=result)
        except Document.DoesNotExist:
            return CustomResponse().base(message="document not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getSubDocument(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            documentId = request.query_params.get('id', None)
            bab = request.query_params.get('bab', None)
            if not documentId:
                return CustomResponse.badRequest('need query param "id"')
            result = []
            if bab:
                data = Bab.objects.filter(
                    bab=bab, detaildocument=documentId)
                serializer = BabSerializer(data, many=True)
                result.extend(serializer.data)
            else:
                data = Bab.objects.filter(detaildocument=documentId)
                serializer = BabSerializer(data, many=True)

                def getSub(x):
                    return {
                        "id": x["id"],
                        "name": x['sub_bab'],
                        "page": x['page'],
                        "end_page": x['endpage'],
                        "file": x['path']
                    }
                for k, v in itertools.groupby(serializer.data, key=lambda x: x['bab']):
                    result.append({
                        "bab": k,
                        "sub_bab": list(map(getSub, v))
                    })
            return CustomResponse.ok(values=result)
        except Bab.DoesNotExist:
            return CustomResponse().base(message="bab not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def searchDoc(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            search = request.query_params.get('search', None)
            if not search:
                return CustomResponse.badRequest('need query param "search"')
            data = Bab.objects.filter(
                text1__icontains=search.replace(" ", ''), detaildocument__status='active').exclude(sub_bab="")

            def cropText(searchText: str, text: str, text1: str):
                totalChar = len(text)-len(text1) + len(search)
                index = text1.lower().index(searchText.lower().replace(' ', ''))
                start = max(index-totalChar, 0)
                end = min(len(text), index+totalChar+len(searchText))
                return text[start:end]
            babSerializer = BabSearchSerializer(data, many=True)
            result = []
            for x in babSerializer.data:
                result.append({
                    "id": x['id'],
                    "bab": x['bab'],
                    "sub_bab": x['sub_bab'],
                    "page": x['page'],
                    "search":  cropText(search, x['text'], x['text1']),
                    "path": x['path'],
                    "detail_doc": x["detaildocument"]
                })
            return CustomResponse.ok(values=result)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def searchDocPaginate(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            search = request.query_params.get('search', None)
            page = request.query_params.get('page', None)

            if not (search and page):
                return CustomResponse.badRequest('need query param "search" & "page"')
            data = Bab.objects.filter(
                text1__icontains=search.replace(" ", ''), detaildocument__status='active').exclude(sub_bab="")

            perPage = 10
            p = Paginator(data, perPage)
            totalPage = p.num_pages
            totalItem = p.count
            try:
                currentPage = p.page(page)
            except PageNotAnInteger:
                # currentPage = paginator.page(1)
                currentPage = p.page(1)
            except EmptyPage:
                # currentPage = paginator.page(totalPage)
                currentPage = p.page(totalPage)

            # def cropText(searchText: str, text: str, text1: str):
            #     totalChar = len(text)-len(text1) + len(search)
            #     index = text1.lower().index(searchText.lower().replace(' ', ''))
            #     start = max(index-totalChar, 0)
            #     end = min(len(text), index+totalChar+len(searchText))
            #     return text[start:end]
            def cropText(searchText: str, text: str, text1: str):
                totalChar = len(text)-len(text1) + len(search)
                index = text.lower().index(searchText.lower())
                if index==0:
                    index = text1.lower().index(searchText.lower().replace(' ', ''))
                    start = max(index-totalChar, 0)
                    end = min(len(text), index+totalChar+len(searchText))
                else:
                    start = max(index-200, 0)
                    end = min(index+200, len(text))
                return text[start:end]

            babSerializer = BabSearchSerializer(
                currentPage, many=True)
            result = []
            for x in babSerializer.data:
                result.append({
                    "id": x['id'],
                    "bab": x['bab'],
                    "sub_bab": x['sub_bab'],
                    "page": x['page'],
                    "search":  cropText(search, x['text'], x['text1']),
                    "path": x['path'],
                    "detail_doc": x["detaildocument"]
                })
            return CustomResponse.ok(values={
                "total_page": totalPage,
                "per_page": perPage,
                "current_page": currentPage.number,
                "count_data": totalItem,
                "data": result
            })
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def searchDocs(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            search = request.query_params.get('search', None)
            if not search:
                return CustomResponse.badRequest('need query param "search"')
            data = Bab.objects.filter(
                text1__icontains=search.replace(" ", ''), detaildocument__status='active').exclude(sub_bab="")

            def cropText(searchText: str, text: str, text1: str):
                listText = []
                for i in re.finditer(search.replace(" ", ''), text1, flags=re.IGNORECASE):
                    totalChar = len(text)-len(text1) + len(search)
                    # index = text1.lower().index(searchText.lower().replace(' ', ''))
                    index = i.start()
                    start = max(index-totalChar, 0)
                    end = min(len(text), index+totalChar+len(searchText))
                    listText.append(text[start: end])
                return listText
            babSerializer = BabSearchSerializer(data, many=True)
            result = []
            for x in babSerializer.data:
                result.append({
                    "id": x['id'],
                    "bab": x['bab'],
                    "sub_bab": x['sub_bab'],
                    "page": x['page'],
                    "search":  cropText(search, x['text'], x['text1']),
                    "path": x['path'],
                    "detail_doc": x["detaildocument"]
                })
            return CustomResponse.ok(values=result)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getDocumentMost(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            sort = request.query_params.get('sort', None)
            if not sort:
                return CustomResponse.badRequest('need query param "sort"')
            elif sort not in ["like", "view"]:
                return CustomResponse.badRequest('sort only accept "like" or "view"')

            if sort == "like":
                # data = DetailDocument.objects.filter(status='active').annotate(
                #     total_likes=Sum('detaildocument_id__likes_count')).order_by('-total_likes')[:5]
                data = DetailDocument.objects.filter(status='active', likes_count__gt=0).order_by('-likes_count')[:5]
            elif sort == "view":
                # data = DetailDocument.objects.filter(status='active').annotate(
                #     total_likes=Sum('detaildocument_id__likes_count')).order_by(
                #     "-views_count")[:5]
                data = DetailDocument.objects.filter(status='active', views_count__gt=0).order_by('-views_count')[:5]

            serializer = DocTotalDetailSerializerNoAnnotate(data, many=True)

            result = []
            for x in serializer.data:
                doc = Document.objects.get(id=x['document'])
                kategori = Kategori.objects.get(id=doc.kategori.id)
                data_like_bab = []
                if sort == "like":
                    bab = Bab.objects.filter(detaildocument=x['id'],likes_count__gt=0).order_by('-likes_count')
                    for dt in bab:
                        dict_like = {}
                        dict_like['bab_id']=dt.id
                        dict_like['bab']=dt.bab
                        dict_like['subbab']=dt.sub_bab
                        dict_like['likes']=dt.likes_count
                        data_like_bab.append(dict_like)
                result.append({
                    "doc_id": x['document'],
                    "id": x['id'],
                    "title": x['title'],
                    "version": x['version'],
                    "id_kategori": kategori.id,
                    "kategori": kategori.name,
                    "like": x['likes_count'],
                    "view": x['views_count'],
                    "create_date": x['create_date'],
                    "bab_sub":data_like_bab
                })

            return CustomResponse.ok(values=result)

        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getDocumentMostUser(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            sort = request.query_params.get('sort', None)
            if not sort:
                return CustomResponse.badRequest('need query param "sort"')
            elif sort not in ["like", "view"]:
                return CustomResponse.badRequest('sort only accept "like" or "view"')

            userid = request.query_params.get('userid', None)
            if sort == "like":
                # data = DetailDocument.objects.filter(status='active').annotate(
                #     total_likes=Sum('detaildocument_id__likes_count')).order_by('-total_likes')[:5]
                data = DetailDocument.objects.filter(status='active', likes_count__gt=0).order_by('-likes_count')#[:5]
            elif sort == "view":
                # data = DetailDocument.objects.filter(status='active').annotate(
                #     total_likes=Sum('detaildocument_id__likes_count')).order_by(
                #     "-views_count")[:5]
                data = DetailDocument.objects.filter(status='active').order_by('-views_count')#[:5]

            data_my =  []
            idx = 0
            for dt in data:
                if idx > 4:
                    break
                if dt.likes:
                    likess = dt.likes.split(',')
                    if userid in likess:
                        data_my.append(dt)
                
            serializer = DocTotalDetailSerializerNoAnnotate(data_my, many=True)

            result = []
            for x in serializer.data:
                doc = Document.objects.get(id=x['document'])
                kategori = Kategori.objects.get(id=doc.kategori.id)
                data_like_bab = []
                if sort == "like":
                    bab = Bab.objects.filter(detaildocument=x['id'],likes_count__gt=0).order_by('-likes_count')
                    for dt in bab:
                        dict_like = {}
                        dict_like['bab_id']=dt.id
                        dict_like['bab']=dt.bab
                        dict_like['subbab']=dt.sub_bab
                        dict_like['likes']=dt.likes_count
                        data_like_bab.append(dict_like)
                result.append({
                    "doc_id": x['document'],
                    "id": x['id'],
                    "title": x['title'],
                    "version": x['version'],
                    "id_kategori": kategori.id,
                    "kategori": kategori.name,
                    "like": x['likes_count'],
                    "view": x['views_count'],
                    "create_date": x['create_date'],
                    "bab_sub":data_like_bab
                })

            return CustomResponse.ok(values=result)

        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)

class Dashboard(viewsets.ModelViewSet):
    def dashboardSuperAdmin(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
           # totalDocs = DetailDocument.objects.filter(status='active').count()
            # totalDocs = DetailDocument.objects.filter(status='active').select_related('document').values('document__kategori_id', 'document_id').annotate(
            #     total=Count('document__kategori_id')).count()
            totalDocs = DetailDocument.objects.filter(status='active').count()
            totalUser = Log.objects.values_list('user')
            # totalLike = Bab.objects.values_list('likes_count')
            totalLike = DetailDocument.objects.filter(status='active').values_list('likes_count')
            
            return CustomResponse.ok(values={
                "document_count": totalDocs,
                "user_active_count": len(list(set(totalUser))),
                "user_nonactive_count": 0,
                "likes_count": sum(sum(totalLike, ())),
            })
        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def dashboardUser(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            id = request.query_params.get('id', None)

            if not id:
                return CustomResponse.badRequest(message="need query param 'id'")

            def getDoc(serializer):
                data = []
                for x in serializer.data:
                    document = Document.objects.get(id=x['document'])
                    serializerDoc = DocSerializer(document)
                    if 'id' in serializerDoc.data.keys():
                        data.append({
                            'date': x['create_date'],
                            "document": serializerDoc.data
                        })
                return data

            recentView = Log.objects.filter(user=id, action="view document").order_by(
                '-create_date')
            recentLike = Log.objects.filter(user=id, action="like document").order_by(
                '-create_date')

            likeSerializer = LogSerializer(recentLike, many=True)
            viewSerializer = LogSerializer(recentView, many=True)

            return CustomResponse.ok(values={
                "likes": getDoc(likeSerializer),
                "views": getDoc(viewSerializer),
            })

        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def countDocs(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            id = request.query_params.get('id', None)
            if id:
                detail = DetailDocument.objects.filter(id=id,status='active')
            else:
                # detail = DetailDocument.objects.all()
                detail = DetailDocument.objects.filter(status='active')
            serializer = DocDetailSerializer(detail, many=True)

            def parseDaftarIsi(text: str):
                res = []
                bab = list(filter(lambda x: x, text.split('|')))
                for x in bab:
                    subBab = x.split('#')
                    res.append({
                        "bab": subBab.pop(0),
                        "list": subBab,
                    })
                return res
            result = []
            for x in serializer.data:
                doc = Document.objects.get(id=x['document'])
                docSerializer = DocSerializer(doc)
                result.append({
                    "doc_id": x['document'],
                    "id": x['id'],
                    'pages': x['pages'],
                    "title": x['title'],
                    "like": docSerializer.data['likes'],
                    "view": docSerializer.data['views'],
                    "file": x['doc_file'],
                    "contents": parseDaftarIsi(x['contents'])
                })
            return CustomResponse.ok(values=result)
        except Document.DoesNotExist:
            return CustomResponse().base(message="document not found", status=404, success=False)
        except Exception as e:
            print(e)
            return CustomResponse().base(message=str(e), status=500, success=False)


class Category(viewsets.ModelViewSet):
    def init(self, request, format=None):

        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            category = Kategori.objects.all()

            if category.count() != 0:
                return CustomResponse().base(status=403, message="kategori already defined")

            categoryData = [
                {
                    "name": "Standar Pemeriksaan Kinerja",
                    "sort": 2,
                    "sub": [
                        {
                            "name": "Pedoman",
                            "sub": [
                                {
                                    "name": "Juklak / Juknis",
                                    "sub": [
                                        {
                                            "name": "Panduan",
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                },
                {
                    "name": "Standar Pemeriksaan Dengan Tujuan Tertentu",
                    "sort": 3,
                    "sub": [
                        {
                            "name": "Pedoman",
                            "sub": [
                                {
                                    "name": "Juklak / Juknis",
                                    "sub": [
                                        {
                                            "name": "Panduan",
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                },
                {
                    "name": "Standar Pemeriksaan Keuangan",
                    "sort": 1,
                    "sub": [
                        {
                            "name": "Pedoman",
                            "sub": [
                                {
                                    "name": "Juklak / Juknis",
                                    "sub": [
                                        {
                                            "name": "Panduan",
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                },
                {
                    "name": "Manajemen Pemeriksaan",
                    "sort": 4,
                    "sub": [
                        {
                            "name": "Pedoman",
                            "sub": [
                                {
                                    "name": "POS Juknis",
                                    "sub": [
                                        {
                                            "name": "IK Panduan",
                                        },
                                        {
                                            "name": "Panduan",
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                },
            ]

            def addData(data, id=None,sort=None):
                serializer = KategoriCreateSerializer(
                    data={"name": data['name'], "parent": id,"sort":sort})
                if serializer.is_valid():
                    serializer.save()
                if "sub" in data.keys():
                    for x in data['sub']:
                        addData(x, serializer.data['id'])

            for x in categoryData:
                addData(x,None,x['sort'])

            return CustomResponse().base(status=201, message="Init Complete")

        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)

    def getAll(self, request, format=None):
        try:
            token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
            ret1, user = authenticate_credentials(token)
            if False == ret1 or None == user:
                return CustomResponse().badRequest(
                    message='token invalid'
                )
            data = Kategori.objects.all().order_by('parent', 'name')
            serializer = KategoriCreateSerializer(data, many=True)
            dataDoc = DetailDocument.objects.filter(status='active').select_related('document').values('document__kategori_id', 'document_id').annotate(
                total=Count('document__kategori_id'))
            from collections import Counter
            vs = list(map(lambda k: k['document__kategori_id'], list(dataDoc)))
            datax = dict(Counter(vs))
            def recursv(objek,total):
                if len(objek['sub'])<1:
                    total+=objek['count']
                    return total
                else:
                    total+=objek['count']
                    return recursv(objek['sub'][0],total)
            def addSub(x):
                return {
                    "id": x['id'],
                    'name': x['name'],
                    'count': datax[x['id']] if x['id'] in datax else 0,
                    'sub': list(map(addSub, filter(lambda y: y['parent'] == x["id"], serializer.data)))
                }
            result = []
            i=0
            for x in filter(lambda x: not x["parent"], serializer.data):
                total=0
                result.append({
                    "id": x['id'],
                    'name': x['name'],
                    'sort': x['sort'],
                    'count': datax[x['id']] if x['id'] in datax else 0,
                    'sub': list(map(addSub, filter(lambda y: y['parent'] == x["id"], serializer.data)))
                })
                final=recursv(result[i],total)
                result[i]['count']=final
                i+=1
            result = sorted(result, key=lambda k: k['sort'])
            return CustomResponse.ok(values=result)
        except Exception as e:
            print("error :", e)
            return CustomResponse().base(message=str(e), status=500, success=False)
